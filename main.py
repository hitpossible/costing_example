import PowerPDF
from datetime import datetime
import calendar
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from collections import defaultdict

thai_months = [
    "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
    "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
]

def convert_to_thai_date_range(date_str):
    parts = date_str.split('/')
    if len(parts) == 2:  # MM/YYYY
        month, year = int(parts[0]), int(parts[1])
        first_day = 1
        last_day = calendar.monthrange(year, month)[1]
        thai_year = year + 543
        thai_month = thai_months[month]
        return f"{first_day} - {last_day} {thai_month} พ.ศ. {thai_year}"
    
    elif len(parts) == 3:  # DD/MM/YYYY
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        thai_year = year + 543
        thai_month = thai_months[month]
        return f"{day} {thai_month} พ.ศ. {thai_year}"
    
    else:
        return "รูปแบบวันที่ไม่ถูกต้อง"
    
def draw_sig_box(pdf, x, y, w, h, name_text, role_text, name_ratio=0.72):
    name_h = h * name_ratio          
    role_h = h - name_h           

    pdf.rect(x, y, w, h)             
    pdf.line(x, y + name_h, x + w, y + name_h) 

    pdf.set_xy(x, y + 7)
    pdf.set_font("TH", "B", 12)
    pdf.cell(w, name_h, name_text, align="C")

    pdf.set_xy(x, y + name_h)
    pdf.set_font("TH", "", 11)
    pdf.cell(w, role_h, role_text, align="C")

def createElectricityReport(data):
    pdf = PowerPDF.PowerPDF(
        title="ค่าไฟฟ้าแยกตามโรงงาน TBKT,TBKK,DIE CASTING,MST",
        period=convert_to_thai_date_range(data["bill_month"]),
        issued=convert_to_thai_date_range(data["issued"]),
    )
    pdf.add_page()

    col_w = [60, 20, 53, 53]
    row_h = 6.5

    pdf.set_font("TH", "B", size=12)
    pdf.set_fill_color(240, 240, 240)

    pdf.cell(col_w[0], row_h, "", border=1, align="C", fill=True)

    right_span_w = col_w[0] + col_w[1]
    left_span_w = col_w[2] + col_w[3]

    pdf.cell(col_w[1] + col_w[2], row_h, "กิโลวัตต์/หน่วย/กิโลวาร์", border=1, align="C", fill=True)
    pdf.cell(col_w[3], row_h, "จำนวนเงิน (บาท)", border=1, align="C", fill=True)
    pdf.ln(row_h)

    energy_total_kw = data["max_peak_kw"] + data["energy_peak1_kw"] + data["energy_offpeak_kw"] + data["energy_holiday1_kw"] + data["energy_peak2_kw"] + data["energy_holiday2_kw"]
    energy_std_baht = data["max_peak_baht"] + data["energy_peak1_baht"] + data["energy_offpeak_baht"] + data["energy_holiday1_baht"] + data["energy_peak2_baht"] + data["energy_holiday2_baht"]
    ft_kw = energy_total_kw - data["max_peak_kw"]
    ft_baht = ft_kw * data["rate_ft"]
    energy_total_baht = energy_std_baht + ft_baht
    all_fac = data["fac_1_kw"] + data["fac_2_kw"] + data["fac_3_kw"] + data["mst_kw"]
    pea_rate = energy_total_baht / energy_total_kw
    tbkk_rate = energy_total_baht / all_fac

    detail = [
        ["ค่าพลังไฟฟ้าสูงสุด (กิโลวัตต์)", "P", data["max_peak_kw"], data["max_peak_baht"]],
        ["พลังงานไฟฟ้า (หน่วย)", "P", data["energy_peak1_kw"], data["energy_peak1_baht"]],
        ["", "OP", data["energy_offpeak_kw"], data["energy_offpeak_baht"]],
        ["", "H",  data["energy_holiday1_kw"], data["energy_holiday1_baht"]],
        ["", "P",  data["energy_peak2_kw"], data["energy_peak2_baht"]],
        ["", "H",  data["energy_holiday2_kw"], data["energy_holiday2_baht"]],
        ["พลังงานไฟฟ้ารวม (หน่วย)", "", energy_total_kw, "-"],
        ["ค่าบริการ", "", "", data["service_charge"]],
        ["ค่าไฟฟ้ามาตรฐาน", "", "", energy_std_baht],
        ["รวมจำนวนเงินค่า Ft (บาท)", data["rate_ft"], ft_kw, ft_baht],
        ["ส่วนลดรัฐบาล", "", "", "-"],
    ]

    pdf.set_font("TH", size=12)

    for desc, period, kwh, amt in detail:
        pdf.cell(col_w[0], row_h, str(desc),   border="LR", align="L")
        pdf.cell(col_w[1], row_h, str(period), border="LR", align="C")
        pdf.cell(col_w[2], row_h, f"{kwh:,.2f}" if isinstance(kwh, (int,float)) else str(kwh), border="LR", align="R")
        pdf.cell(col_w[3], row_h, f"{amt:,.2f}" if isinstance(amt, (int,float)) else str(amt), border="LR", align="R")
        pdf.ln(row_h)

    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "รวมค่าพลังงานไฟฟ้า", border="TLR", align="L")
    pdf.set_font("TH", "B", size=12)
    pdf.cell(col_w[3], row_h, f"{energy_total_baht:,.2f}", border="TLR", align="R")
    pdf.ln(row_h)
    pdf.set_font("TH", size=12)
    pdf.cell(col_w[0] + col_w[1], row_h, "ค่าน้ำหนักคิดค่าไฟ (บาท/กิโลวัตต์)", border="LB", align="L")
    pdf.cell(col_w[2], row_h, f"{pea_rate:,.2f}", border="BR", align="R")
    pdf.cell(col_w[3], row_h, f"{tbkk_rate:,.2f}", border="BR", align="R")
    pdf.ln(row_h * 2)


    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("TH", "B", size=12)
    pdf.cell(right_span_w + left_span_w, row_h, "ค่าตัวเลขมิเตอร์ไฟฟ้าแยกแต่ละโรงงาน", border=1, align="C", fill=True)
    pdf.ln(row_h)

    detail = [
        ["F10-1", data["fac_1_kw"]],
        ["F10-2", data["fac_2_kw"]],
        ["F10-3", data["fac_3_kw"]],
        ["MST", data["mst_kw"]],
    ]
    pdf.set_font("TH", size=12)
    for fac_name, amt in detail:
        border_style = "LRT" if fac_name == "MST" else "LR"

        pdf.cell(left_span_w, row_h, str(fac_name), border=border_style, align="L")
        pdf.cell(right_span_w, row_h, f"{amt:,.2f}", border=border_style, align="R")
        pdf.ln(row_h)


    meter_err = ((1-(all_fac/energy_total_kw)))

    pdf.cell(left_span_w, row_h, "พลังงานไฟฟ้ารวม (หน่วย)", border="LRT", align="L")
    pdf.cell(right_span_w, row_h, f"{all_fac:,.2f}", border="LRT", align="R")
    pdf.ln(row_h)
    pdf.cell(left_span_w, row_h, "ค่าความผิดพลาดของมิเตอร์ % (ERR)", border="LRB", align="L")
    pdf.cell(right_span_w, row_h, f"({meter_err:.2%})", border="LRB", align="R")
    pdf.ln(row_h)
    pdf.ln(row_h)

    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("TH", "B", size=12)
    pdf.cell(right_span_w + left_span_w, row_h, "ค่าไฟฟ้าจากการคำนวณ ไม่รวม VAT 7%", border=1, align="C", fill=True)
    pdf.ln(row_h)

    tbkk_total_amount = data["direct_kw"] + data["admin_kw"] + data["indirect_kw"]
    tbkk_total_baht = data["direct_baht"] + data["admin_baht"] + data["indirect_baht"]
    all_fac_total_amount = tbkk_total_amount + data["mst_kw"]
    all_fac_total_baht = tbkk_total_baht + data["mst_baht"]

    pdf.set_font("TH", size=12)
    pdf.cell(col_w[0] + col_w[1], row_h, "1.TBKK (MW)X1000*ERR", border="LR", align="L")
    pdf.cell(col_w[2], row_h, f"{tbkk_total_amount:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{tbkk_total_baht:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "DIRECT TBKK", border="LR", align="R")
    pdf.cell(col_w[2], row_h, f"{data['direct_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['direct_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "ADMIN TBKK", border="LR", align="R")
    pdf.cell(col_w[2], row_h, f"{data['admin_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['admin_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "INDIRECT TBKK", border="LR", align="R")
    pdf.cell(col_w[2], row_h, f"{data['indirect_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['indirect_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "2.MST : Advance (MW)X1000*ERR", border="LRT", align="L")
    pdf.cell(col_w[2], row_h, f"{data['mst_kw']:,.2f}", border="LRT", align="R")
    pdf.cell(col_w[3], row_h, f"{data['mst_baht']:,.2f}", border="LRT", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "พลังงานไฟฟ้ารวม (หน่วย)", border="LRT", align="L")
    pdf.cell(col_w[2], row_h, f"{all_fac_total_amount:,.2f}", border="LRT", align="R")
    pdf.cell(col_w[3], row_h, f"-", border="LRT", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1], row_h, "ค่าไฟฟ้ามาตรฐาน", border="LRB", align="L")
    pdf.cell(col_w[2], row_h, f"-", border="LRB", align="R")
    pdf.cell(col_w[3], row_h, f"{all_fac_total_baht:,.2f}", border="LRB", align="R")
    pdf.ln(row_h)

    y = 255
    h = 30

    content_w = pdf.w - pdf.l_margin - pdf.r_margin
    gap = 40.0                      
    group_w = (content_w - gap) / 2
    small_w = group_w / 2
    big_w = group_w            

    x_left = pdf.l_margin
    draw_sig_box(pdf, x_left,             y, small_w, h, "Mr.Suthad T.",   "Approved By")
    draw_sig_box(pdf, x_left + small_w,   y, small_w, h, "Mr.Atsadang R.", "Approved By")

    x_right = pdf.l_margin + group_w + gap
    draw_sig_box(pdf, x_right, y, big_w, h, "Mrs.Wandee P.", "Accounting Section")

    bill_month = data["bill_month"] 
    month, year = bill_month.split("/")

    folder_path = os.path.join("reports", "electric", year)
    os.makedirs(folder_path, exist_ok=True)

    file_name = f"electric_report_{int(month):02d}_{year}.pdf"
    file_path = os.path.join(folder_path, file_name)

    pdf.output(file_path)

def createSolarReport(data):
    pdf = PowerPDF.PowerPDF(
        title="ค่าไฟฟ้าโซล่าเซลล์",
        period=convert_to_thai_date_range(data["bill_month"]),
        issued=convert_to_thai_date_range(data["issued"]),
    )
    pdf.add_page()

    col_w = [60, 8, 40, 40, 40]
    row_h = 6.5

    pdf.set_font("TH", "B", size=12)
    pdf.set_fill_color(240, 240, 240)
    pdf.ln(8)

    pdf.cell(col_w[0] + col_w[1], row_h, "รายการ", border=1, align="C", fill=True)
    pdf.cell(col_w[2], row_h, "กิโลวัตต์/หน่วย/กิโลวาร์", border=1, align="C", fill=True)
    pdf.cell(col_w[3], row_h, "จำนวนเงินก่อนหักส่่วนลด (บาท)", border=1, align="C", fill=True)
    pdf.cell(col_w[3], row_h, "จำนวนเงินหลังหักส่่วนลด (บาท)", border=1, align="C", fill=True)
    pdf.ln(row_h)

    pdf.set_font("TH", size=12)

    pdf.cell(col_w[0], row_h, "ค่าพลังงานไฟฟ้าช่วง Peak", border="LR", align="L")
    pdf.cell(col_w[1], row_h, "P", border="LR", align="C")
    pdf.cell(col_w[2], row_h, f"{data['power_peak_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['power_peak_before_discount']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['power_peak_after_discount']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)

    pdf.cell(col_w[0], row_h, "ค่าพลังงานไฟฟ้าช่วง Off Peak", border="LR", align="L")
    pdf.cell(col_w[1], row_h, "OP", border="LR", align="C")
    pdf.cell(col_w[2], row_h, f"{data['power_offpeak_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['power_offpeak_before_discount']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['power_offpeak_after_discount']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)

    pdf.cell(col_w[0], row_h, "ค่าพลังงานไฟฟ้าช่วง Holiday", border="LRB", align="L")
    pdf.cell(col_w[1], row_h, "H", border="LRB", align="C")
    pdf.cell(col_w[2], row_h, f"{data['power_holiday_kw']:,.2f}", border="LRB", align="R")
    pdf.cell(col_w[3], row_h, f"{data['power_holiday_before_discount']:,.2f}", border="LRB", align="R")
    pdf.cell(col_w[4], row_h, f"{data['power_holiday_after_discount']:,.2f}", border="LRB", align="R")
    pdf.ln(row_h)

    pdf.cell(col_w[0] + col_w[1], row_h, "ค่าไฟฟ้าผันแปร (Ft)", border="LR", align="L")
    pdf.cell(col_w[2], row_h, f"", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['ft_before_discount']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['ft_after_discount']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)

    power_total = data['power_peak_kw'] + data['power_offpeak_kw'] + data['power_holiday_kw']
    pdf.cell(col_w[0] + col_w[1], row_h, "พลังงานไฟฟ้ารวม (หน่วย)", border="LR", align="L")
    pdf.cell(col_w[2], row_h, f"{power_total:,.2f}", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"", border="LR", align="R")
    pdf.ln(row_h)

    pdf.cell(col_w[0] + col_w[1], row_h, "ค่าความต้องการพลังไฟฟ้า (บาท)", border="LR", align="L")
    pdf.cell(col_w[2], row_h, f"", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['power_demand']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)

    total_power_before_discount = data["power_peak_before_discount"] + data["power_offpeak_before_discount"] + data["power_holiday_before_discount"] + data["ft_before_discount"]
    total_power_after_discount = data["power_peak_after_discount"] + data["power_offpeak_after_discount"] + data["power_holiday_after_discount"] + data["ft_after_discount"] + data["power_demand"]
    pdf.cell(col_w[0] + col_w[1], row_h, "รวมค่าพลังงานไฟฟ้า", border="LRB", align="L")
    pdf.cell(col_w[2], row_h, f"", border="LRB", align="R")
    pdf.cell(col_w[3], row_h, f"{total_power_before_discount:,.2f}", border="LRB", align="R")
    pdf.set_font("TH", "B", size=12)
    pdf.cell(col_w[4], row_h, f"{total_power_after_discount:,.2f}", border="LRB", align="R")
    pdf.ln(row_h)

    pdf.set_font("TH", size=12)

    rate_before_discount = total_power_before_discount / power_total
    rate_after_discount = total_power_after_discount / power_total
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "ค่าน้ำหนักคิดค่าไฟ (บาท/กิโลวัตต์)", border="LRB", align="L")
    pdf.cell(col_w[3], row_h, f"{rate_before_discount:,.2f}", border="LRB", align="R")
    pdf.cell(col_w[4], row_h, f"{rate_after_discount:,.2f}", border="LRB", align="R")
    pdf.ln(row_h)
    pdf.ln(row_h)

    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("TH", "B", size=12)
    pdf.cell(col_w[0] + col_w[1] + col_w[2] + col_w[3] + col_w[4], row_h, "ค่าไฟฟ้าจากการคำนวณ ไม่รวม VAT 7%", border=1, align="C", fill=True)
    pdf.ln(row_h)

    tbkk_total_amount = data["direct_kw"] + data["admin_kw"] + data["indirect_kw"]
    tbkk_total_baht = data["direct_baht"] + data["admin_baht"] + data["indirect_baht"]

    pdf.set_font("TH", size=12)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "1.TBKK (MW)X1000*ERR", border="LR", align="L")
    pdf.cell(col_w[3], row_h, f"{tbkk_total_amount:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{tbkk_total_baht:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "DIRECT TBKK", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['direct_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['direct_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "ADMIN TBKK", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['admin_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['admin_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "INDIRECT TBKK", border="LR", align="R")
    pdf.cell(col_w[3], row_h, f"{data['indirect_kw']:,.2f}", border="LR", align="R")
    pdf.cell(col_w[4], row_h, f"{data['indirect_baht']:,.2f}", border="LR", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "พลังงานไฟฟ้ารวม (หน่วย)", border="LRT", align="L")
    pdf.cell(col_w[3], row_h, f"{tbkk_total_amount:,.2f}", border="LRT", align="R")
    pdf.cell(col_w[4], row_h, f"-", border="LRT", align="R")
    pdf.ln(row_h)
    pdf.cell(col_w[0] + col_w[1] + col_w[2], row_h, "ค่าไฟฟ้ามาตรฐาน", border="LRB", align="L")
    pdf.cell(col_w[3], row_h, f"-", border="LRB", align="R")
    pdf.cell(col_w[4], row_h, f"{tbkk_total_baht:,.2f}", border="LRB", align="R")
    pdf.ln(row_h)

    y = 170
    h = 30

    content_w = pdf.w - pdf.l_margin - pdf.r_margin
    gap = 40.0                      
    group_w = (content_w - gap) / 2
    small_w = group_w / 2
    big_w = group_w            

    x_left = pdf.l_margin
    draw_sig_box(pdf, x_left,             y, small_w, h, "Mr.Watcharagorn K.",   "Prepared By")
    draw_sig_box(pdf, x_left + small_w,   y, small_w, h, "Mr.Atsadang R.", "Approved By")

    x_right = pdf.l_margin + group_w + gap
    draw_sig_box(pdf, x_right, y, big_w, h, "Mrs.Wandee P.", "Accounting Section")


    bill_month = data["bill_month"] 
    month, year = bill_month.split("/")

    folder_path = os.path.join("reports", "solar", year)
    os.makedirs(folder_path, exist_ok=True)

    file_name = f"solar_report_{int(month):02d}_{year}.pdf"
    file_path = os.path.join(folder_path, file_name)

    pdf.output(file_path)


def createExcelReport(data, filename="report.xlsx"):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "All Data"

    # ===== style =====
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ----- ตารางรวมทั้งหมด -----
    cols_all = [
        ("GROUP PD",       "department"),
        ("SECTION",        "section"),
        ("LINE",           "line"),
        ("PRODUCTION TIME","production_time"),
        ("KW (PE)",        "kw"),
        ("KWH",            "kwh"),
        ("KWH (UT)",       "kwh_ut"),
        ("AMOUNT",         "amount"),
        ("AMOUNT SOLAR",   "amount_solar"),
        ("TOTAL AMOUNT",   "total_amount"),
    ]

    ws1.append([c[0] for c in cols_all])

    # apply style ให้หัวตาราง
    for col_num, (header, _) in enumerate(cols_all, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # data rows
    for row in data:
        values = [row.get(c[1], "") for c in cols_all]
        ws1.append(values)

    # apply border ให้ทุก cell
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(cols_all)):
        for cell in row:
            cell.border = thin_border

    # autosize
    for col in ws1.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # ----- ตารางสรุปตาม department -----
    ws2 = wb.create_sheet("By Department")

    grouped = defaultdict(lambda: {"kwh_ut":0, "amount":0, "amount_solar":0, "total_amount":0})
    for row in data:
        dep = row.get("department", "")
        grouped[dep]["kwh_ut"]       += float(row.get("kwh_ut", 0) or 0)
        grouped[dep]["amount"]       += float(row.get("amount", 0) or 0)
        grouped[dep]["amount_solar"] += float(row.get("amount_solar", 0) or 0)
        grouped[dep]["total_amount"] += float(row.get("total_amount", 0) or 0)

    cols_dep = [
        ("DEPARTMENT",    "department"),
        ("KWH (UT)",      "kwh_ut"),
        ("AMOUNT",        "amount"),
        ("AMOUNT SOLAR",  "amount_solar"),
        ("TOTAL AMOUNT",  "total_amount"),
    ]

    ws2.append([c[0] for c in cols_dep])
    for col_num, (header, _) in enumerate(cols_dep, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for dep, vals in grouped.items():
        ws2.append([
            dep,
            vals["kwh_ut"],
            vals["amount"],
            vals["amount_solar"],
            vals["total_amount"],
        ])

    # apply border
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(cols_dep)):
        for cell in row:
            cell.border = thin_border

    # autosize
    for col in ws2.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(filename)

if __name__ == "__main__":
    data = {
        "bill_month": "8/2025",
        "issued": "2/9/2025",
        "max_peak_kw": 2850,
        "max_peak_baht": 211299.00,
        "energy_peak1_kw": 497100,
        "energy_peak1_baht": 2038952.78,
        "energy_offpeak_kw": 423600,
        "energy_offpeak_baht": 1683064.58,
        "energy_holiday1_kw": 232200,
        "energy_holiday1_baht": 0.00,
        "energy_peak2_kw": 0,
        "energy_peak2_baht": 0.00,
        "energy_holiday2_kw": 0,
        "energy_holiday2_baht": 0.00,
        "service_charge": 312.24,
        "rate_ft": 0.1972,
        "fac_1_kw": 1000,
        "fac_2_kw": 200,
        "fac_3_kw": 300,
        "mst_kw": 400,
        "direct_kw": 1000,
        "admin_kw": 200,
        "indirect_kw": 300,
        "direct_baht": 30000,
        "indirect_baht": 30000,
        "admin_baht": 30000,
        "mst_kw": 400,
        "mst_baht": 30000,

    }
    createElectricityReport(data)


    solar_data = {
        "bill_month": "8/2025",
        "issued": "2/9/2025",
        "power_peak_kw": 85655.98,
        "power_peak_before_discount": 351403.66,
        "power_peak_after_discount": 182729.91,
        "power_offpeak_kw": 65347.66,
        "power_offpeak_before_discount": 168917.17,
        "power_offpeak_after_discount": 87836.93,
        "power_holiday_kw": 67064.98,
        "power_holiday_before_discount": 173356.26,
        "power_holiday_after_discount": 90145.26,
        "ft_before_discount":  43003.13,
        "ft_after_discount": 22361.63,
        "power_demand": 63360.00,
        "direct_kw": 1082906.10,
        "admin_kw": 38287.40,
        "indirect_kw": 34526.90,
        "direct_baht": 3931725.64,
        "admin_baht": 139010.72,
        "indirect_baht": 125357.41,
    }
    createSolarReport(solar_data)
    

    excel_data = data = [
        {
            "department": "K1PD01",
            "section": "K1A000",
            "line": "K1A001",
            "production_time": 100,
            "kw": 12,
            "kwh": 2100,
            "kwh_ut": 3000,
            "amount": 40000,
            "amount_solar": 10000,
            "total_amount": 50000
        },
        {
            "department": "K1PD01",
            "section": "K1A000",
            "line": "K1A002",
            "production_time": 80,
            "kw": 10,
            "kwh": 1800,
            "kwh_ut": 2500,
            "amount": 30000,
            "amount_solar": 8000,
            "total_amount": 38000
        },
        {
            "department": "K2PD02",
            "section": "K2A000",
            "line": "K2A001",
            "production_time": 120,
            "kw": 15,
            "kwh": 2500,
            "kwh_ut": 3300,
            "amount": 50000,
            "amount_solar": 12000,
            "total_amount": 62000
        }
    ]

    createExcelReport(excel_data)
